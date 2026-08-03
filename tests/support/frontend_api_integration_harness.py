"""Bridge runtime browser FormData into the real Python production stack.

The Vitest caller supplies fields captured from ``useUploadBlastCsv``.  This
module deliberately does not define or translate geometry field names.
"""

from __future__ import annotations

import io
import json
import os
from pathlib import Path
import sys
import tempfile


REPOSITORY_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPOSITORY_ROOT))


def main() -> None:
    request = json.load(sys.stdin)
    with tempfile.TemporaryDirectory(prefix="blast-browser-api-") as data_dir:
        os.environ["CONCILIACION_DATA_DIR"] = data_dir

        from fastapi.testclient import TestClient
        from openpyxl import load_workbook

        import api.database as db
        from api.main import app

        db.init_db()
        session_id = db.create_session()
        fields = dict(request["fields"])
        fields["session_id"] = session_id

        with TestClient(app) as client:
            response = client.post(
                "/api/v1/blast/upload",
                data=fields,
                files={
                    "file": (
                        "pozos.csv",
                        io.BytesIO(request["csv"].encode("utf-8")),
                        "text/csv",
                    )
                },
            )
            body = response.json()
            if response.status_code != 200:
                raise RuntimeError(f"upload failed: {response.status_code} {body}")

            persisted = db.get_settings(session_id)
            exported = client.get(
                "/api/v1/export/blast-diagnostics",
                headers={"X-Session-ID": session_id},
            )
            exported.raise_for_status()

        workbook = load_workbook(io.BytesIO(exported.content), read_only=True)
        expected_sheets = {
            "Pozos_Aceptados",
            "Filas_Rechazadas",
            "Advertencias",
            "Errores_Bloqueantes",
            "Resumen_Procesamiento",
            "Configuracion_Geometrica",
            "Diagnostico_Espacial",
        }
        result = {
            "status_code": response.status_code,
            "accepted_rows": len(body["accepted_rows"]),
            "persisted_same_result": (
                persisted["accepted_rows"] == body["accepted_rows"]
                and persisted["event_warnings"] == body["event_warnings"]
                and persisted["spatial_diagnostics"] == body["spatial_diagnostics"]
            ),
            "export_reopened": expected_sheets.issubset(workbook.sheetnames),
            "inclination_unit": body["geometry_configuration"]["inclination_unit"],
            "azimuth_unit": body["geometry_configuration"]["azimuth_unit"],
        }
        workbook.close()
        json.dump(result, sys.stdout)


if __name__ == "__main__":
    main()
