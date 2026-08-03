"""Tests for the structured blast-processing diagnostics export.

Integración §3.7/5.7 — the export MUST:
- exist for both full diagnostics and rejected-only,
- work with zero accepted rows,
- preserve every rejection field,
- preserve the geometry configuration,
- produce a file that can be re-opened by openpyxl/pandas.
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest

from core.blast_export import (
    export_processing_diagnostics_excel,
    export_rejected_rows_excel,
    read_back_excel,
)
from core.geometry_contract import GEOMETRY_CONFIGURATION_VERSION


def _full_payload(*, n_accepted: int = 2, n_rejected: int = 1) -> dict:
    """Build a representative payload mirroring the API processor output."""
    return {
        "accepted_rows": [
            {
                "hole_id": f"H-{i}",
                "X": 1000.0 + i,
                "Y": 2000.0 + i,
                "Z_collar": 4015.0,
                "X_toe": 1003.1 + i,
                "Y_toe": 2000.0 + i,
                "Z_toe": 4003.4,
                "Incl": 15.0,
                "Az": 90.0,
                "inclination_normalized_from_vertical_deg": 15.0,
                "azimuth_normalized_clockwise_from_north_deg": 90.0,
                "row_processing_status": "accepted",
                "geometry_configuration_version": GEOMETRY_CONFIGURATION_VERSION,
            }
            for i in range(n_accepted)
        ],
        "rejected_rows": [
            {
                "hole_id": f"BAD-{i}",
                "source_row_index": 100 + i,
                "source_column": "Latitud_Geo",
                "original_value": None,
                "error_code": "INVALID_X",
                "rejection_reason": "valor no numérico o ausente",
                "affected_calculations": "toe, PF, geometría dependiente",
                "recommended_action": "Corrija el dato original y reprocese.",
                "row_processing_status": "rejected",
            }
            for i in range(n_rejected)
        ],
        "event_warnings": [
            {"code": "INCL_OUT_OF_RANGE", "message": "inclinación > 90° en 1 fila"}
        ],
        "blocking_errors": [],
        "processing_summary": {
            "rows_received": 3,
            "rows_accepted": n_accepted,
            "rows_rejected": n_rejected,
            "geometry_configuration_version": GEOMETRY_CONFIGURATION_VERSION,
        },
        "geometry_configuration": {
            "geometry_configuration_version": GEOMETRY_CONFIGURATION_VERSION,
            "geometry_user_confirmed": True,
            "inclination_convention": "FROM_VERTICAL",
            "inclination_unit": "DEGREES",
            "azimuth_convention": "CLOCKWISE_FROM_NORTH",
            "azimuth_unit": "DEGREES",
            "inclination_source_column": "Inclinacion_real",
            "azimuth_source_column": "Azimuth_real",
        },
        "spatial_diagnostics": {"domain_area_m2": 300.0, "n_interior": 2},
    }


class TestProcessingDiagnosticsExcel:
    def test_full_export_has_all_sheets(self, tmp_path: Path):
        out = export_processing_diagnostics_excel(_full_payload(), tmp_path / "diag.xlsx")
        assert out.exists()
        sheets = read_back_excel(out)
        for name in (
            "Pozos_Aceptados",
            "Filas_Rechazadas",
            "Advertencias",
            "Errores_Bloqueantes",
            "Resumen_Procesamiento",
            "Configuracion_Geometrica",
            "Diagnostico_Espacial",
        ):
            assert name in sheets, f"missing sheet {name}"

    def test_accepted_rows_survive_roundtrip(self, tmp_path: Path):
        out = export_processing_diagnostics_excel(_full_payload(n_accepted=3), tmp_path / "a.xlsx")
        df = read_back_excel(out)["Pozos_Aceptados"]
        assert len(df) == 3
        assert set(["hole_id", "X", "Y", "Z_toe"]).issubset(df.columns)
        assert df["hole_id"].iloc[0] == "H-0"
        # Excel may coerce numeric-looking strings; compare loosely.
        assert str(df["geometry_configuration_version"].iloc[0]) == GEOMETRY_CONFIGURATION_VERSION

    def test_rejected_rows_preserve_all_fields(self, tmp_path: Path):
        out = export_processing_diagnostics_excel(_full_payload(n_rejected=2), tmp_path / "r.xlsx")
        df = read_back_excel(out)["Filas_Rechazadas"]
        assert len(df) == 2
        required = {
            "hole_id",
            "source_row_index",
            "source_column",
            "original_value",
            "error_code",
            "rejection_reason",
            "affected_calculations",
            "recommended_action",
            "row_processing_status",
        }
        assert required.issubset(set(df.columns)), (
            f"missing fields: {required - set(df.columns)}"
        )
        assert df["error_code"].iloc[0] == "INVALID_X"

    def test_zero_accepted_rows_export_works(self, tmp_path: Path):
        payload = _full_payload(n_accepted=0, n_rejected=3)
        payload["blocking_errors"] = [
            {
                "error_code": "NO_ACCEPTED_ROWS",
                "message": "Ninguna fila pasó la validación.",
            }
        ]
        out = export_processing_diagnostics_excel(payload, tmp_path / "zero.xlsx")
        sheets = read_back_excel(out)
        # Pozos_Aceptados may be empty/placeholder; Filas_Rechazadas MUST keep all 3.
        df_rej = sheets["Filas_Rechazadas"]
        assert len(df_rej) == 3
        # Blocking errors sheet has 1 row.
        df_err = sheets["Errores_Bloqueantes"]
        assert len(df_err) == 1
        assert df_err["error_code"].iloc[0] == "NO_ACCEPTED_ROWS"

    def test_geometry_configuration_exported_with_version(self, tmp_path: Path):
        out = export_processing_diagnostics_excel(_full_payload(), tmp_path / "cfg.xlsx")
        df = read_back_excel(out)["Configuracion_Geometrica"]
        # _write_kv_sheet puts key in col A and value in col B
        # read_back_excel returns first row as header — so the kv pairs
        # become rows under unnamed columns. Find the version row by
        # scanning the values.
        version_found = False
        for _, row in df.iterrows():
            for v in row.values:
                if str(v) == "geometry_configuration_version":
                    version_found = True
                    break
        assert version_found, "geometry_configuration_version key not found in export"

    def test_multiple_errors_per_row_preserved(self, tmp_path: Path):
        payload = _full_payload(n_accepted=0, n_rejected=0)
        # Two rejection records for the SAME source row → two rows in the sheet.
        payload["rejected_rows"] = [
            {
                "hole_id": "DUP-1",
                "source_row_index": 5,
                "source_column": "Latitud_Geo",
                "error_code": "INVALID_X",
                "rejection_reason": "x nan",
                "row_processing_status": "rejected",
            },
            {
                "hole_id": "DUP-1",
                "source_row_index": 5,
                "source_column": "Longitud_Geo",
                "error_code": "INVALID_Y",
                "rejection_reason": "y nan",
                "row_processing_status": "rejected",
            },
        ]
        out = export_processing_diagnostics_excel(payload, tmp_path / "multi.xlsx")
        df = read_back_excel(out)["Filas_Rechazadas"]
        assert len(df) == 2
        assert set(df["source_column"]) == {"Latitud_Geo", "Longitud_Geo"}


class TestRejectedOnlyExport:
    def test_standalone_export_works(self, tmp_path: Path):
        payload = _full_payload(n_accepted=0, n_rejected=4)
        out = export_rejected_rows_excel(
            payload["rejected_rows"],
            tmp_path / "only_rej.xlsx",
            config=payload["geometry_configuration"],
        )
        assert out.exists()
        sheets = read_back_excel(out)
        assert "Filas_Rechazadas" in sheets
        assert len(sheets["Filas_Rechazadas"]) == 4
        # Metadata sheet echoes the contract version.
        assert "Metadata" in sheets

    def test_export_reopenable_by_openpyxl(self, tmp_path: Path):
        from openpyxl import load_workbook

        out = export_processing_diagnostics_excel(_full_payload(), tmp_path / "wb.xlsx")
        # Must not raise.
        wb = load_workbook(str(out))
        names = wb.sheetnames
        assert "Pozos_Aceptados" in names
        assert "Filas_Rechazadas" in names
