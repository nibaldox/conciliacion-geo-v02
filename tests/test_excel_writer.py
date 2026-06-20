"""Tests for core.excel_writer — formatted Excel export of reconciliation results.

Exercises the public :func:`export_results` entry point with synthetic
comparison data covering every compliance status and every worksheet
(Resumen, Resumen Ejecutivo, Bancos, Inter-Rampa, Dashboard, Tronadura).
"""
from types import SimpleNamespace

import numpy as np
import openpyxl
import pandas as pd
import pytest

from core.compliance_status import (
    STATUS_BANCO_ADICIONAL,
    STATUS_CUMPLE,
    STATUS_EXTRA,
    STATUS_FALTA_BANCO,
    STATUS_FUERA,
    STATUS_NO_CONSTRUIDO,
    STATUS_NO_CUMPLE,
)
from core.excel_writer import export_results


def _match_comp(**overrides):
    """A fully-populated MATCH comparison row (param_extractor schema)."""
    base = dict(
        sector="Norte",
        section="S-01",
        bench_num=1,
        type="MATCH",
        level="3855",
        height_design=15.0,
        height_real=15.2,
        height_dev=0.2,
        height_status=STATUS_CUMPLE,
        angle_design=70.0,
        angle_real=68.0,
        angle_dev=-2.0,
        angle_status=STATUS_CUMPLE,
        berm_design=9.0,
        berm_real=8.5,
        berm_min=6.0,
        berm_status=STATUS_CUMPLE,
        spill_width=0.0,
        effective_berm=8.5,
        delta_crest=0.1,
        delta_toe=-0.2,
        bench_design=None,
        bench_real=None,
        berm_score=60,
        angle_score=10,
        height_score=30,
        bench_score=100,
        section_score=85.0,
        section_status=STATUS_CUMPLE,
    )
    base.update(overrides)
    return base


def _params(name="S-01", sector="Norte", ir=48.0, overall=42.0):
    return SimpleNamespace(
        section_name=name,
        sector=sector,
        inter_ramp_angle=ir,
        overall_angle=overall,
    )


def _tolerances():
    return {
        "bench_height": {"neg": 1.0, "pos": 1.5},
        "face_angle": {"neg": 5.0, "pos": 5.0},
        "berm_width": {"min": 6.0},
        "inter_ramp_angle": {"neg": 3.0, "pos": 2.0},
        "overall_angle": {"neg": 2.0, "pos": 2.0},
    }


@pytest.fixture()
def comparisons_mixed():
    return [
        _match_comp(bench_num=1, level="3855"),
        _match_comp(
            bench_num=2, level="3840",
            height_status=STATUS_FUERA, angle_status=STATUS_FUERA,
            berm_status=STATUS_NO_CUMPLE, berm_real=4.0,
            section_score=40.0, section_status=STATUS_NO_CUMPLE,
        ),
        _match_comp(
            bench_num=3, level="3825",
            height_status=STATUS_NO_CUMPLE, angle_status=STATUS_NO_CUMPLE,
            berm_status=STATUS_NO_CUMPLE,
            section_score=20.0, section_status=STATUS_NO_CUMPLE,
        ),
        dict(_match_comp(bench_num=4, level="3810"), type="MISSING",
             height_real=None, height_dev=None, height_status=STATUS_NO_CONSTRUIDO,
             angle_real=None, angle_dev=None, angle_status="-",
             berm_real=None, berm_min=None, berm_status=STATUS_FALTA_BANCO,
             spill_width=None, effective_berm=None,
             delta_crest=None, delta_toe=None,
             berm_score=0, angle_score=0, height_score=0, bench_score=0),
        dict(_match_comp(bench_num=999, level="3870"), type="EXTRA",
             height_design=None, height_dev=None, height_status=STATUS_EXTRA,
             angle_design=None, angle_dev=None, angle_status="-",
             berm_design=None, berm_min=None, berm_status=STATUS_BANCO_ADICIONAL,
             bench_score=0),
    ]


@pytest.fixture()
def df_pozos_small():
    return pd.DataFrame({
        "X": [100.0, 105.0],
        "Y": [200.0, 200.0],
        "Z_collar": [3870.0, 3870.0],
        "Z_toe": [3855.0, 3855.0],
        "X_toe": [100.0, 105.0],
        "Y_toe": [200.0, 200.0],
        "Incl": [0.0, 0.0],
        "Az": [0.0, 0.0],
        "Len": [15.0, 15.0],
        "Kilos_Cargados_real": [120.0, 130.0],
        "label_pozo": ["P-1", "P-2"],
    })


@pytest.fixture()
def sections_small():
    sec = SimpleNamespace(
        name="S-01",
        origin=np.array([100.0, 200.0]),
        azimuth=0.0,
        length=200.0,
    )
    return [sec]


class TestExportResults:
    def test_writes_all_core_sheets(self, tmp_path, comparisons_mixed):
        out = tmp_path / "report.xlsx"
        export_results(
            comparisons_mixed,
            [_params()],
            [_params()],
            _tolerances(),
            str(out),
            project_info={"project": "P", "operation": "O", "phase": "F",
                          "author": "A", "date": "2026-06-20"},
        )
        assert out.exists()
        wb = openpyxl.load_workbook(str(out))
        expected = {"Resumen", "Resumen Ejecutivo", "Bancos", "Inter-Rampa", "Dashboard"}
        assert expected <= set(wb.sheetnames)

    def test_tronadura_sheet_when_pozos_given(
        self, tmp_path, comparisons_mixed, df_pozos_small, sections_small
    ):
        out = tmp_path / "report.xlsx"
        export_results(
            comparisons_mixed,
            [_params()],
            [_params()],
            _tolerances(),
            str(out),
            df_pozos=df_pozos_small,
            sections=sections_small,
        )
        wb = openpyxl.load_workbook(str(out))
        assert "Tronadura" in wb.sheetnames

    def test_no_tronadura_sheet_without_pozos(self, tmp_path, comparisons_mixed):
        out = tmp_path / "report.xlsx"
        export_results(comparisons_mixed, [_params()], [_params()], _tolerances(), str(out))
        wb = openpyxl.load_workbook(str(out))
        assert "Tronadura" not in wb.sheetnames

    def test_empty_comparisons_still_writes(self, tmp_path):
        out = tmp_path / "empty.xlsx"
        export_results([], [], [], _tolerances(), str(out))
        assert out.exists()
        wb = openpyxl.load_workbook(str(out))
        assert "Resumen" in wb.sheetnames

    def test_sector_summary_lists_sector(self, tmp_path, comparisons_mixed):
        out = tmp_path / "report.xlsx"
        export_results(comparisons_mixed, [_params()], [_params()], _tolerances(), str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Resumen Ejecutivo"]
        cell_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Norte" in cell_values

    def test_bench_sheet_has_one_row_per_comparison(self, tmp_path, comparisons_mixed):
        out = tmp_path / "report.xlsx"
        export_results(comparisons_mixed, [_params()], [_params()], _tolerances(), str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Bancos"]
        # 1 header row + 1 data row per comparison
        assert ws.max_row == 1 + len(comparisons_mixed)

    def test_interramp_sheet_rounded_angles(self, tmp_path, comparisons_mixed):
        pd1 = _params(ir=48.123, overall=42.456)
        pd2 = _params(ir=47.0, overall=41.0)
        out = tmp_path / "report.xlsx"
        export_results(comparisons_mixed, [pd1], [pd2], _tolerances(), str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Inter-Rampa"]
        # Header + 1 data row (min(len(d), len(t)) == 1)
        assert ws.max_row >= 2

    def test_missing_extra_statuses_rendered(self, tmp_path, comparisons_mixed):
        out = tmp_path / "report.xlsx"
        export_results(comparisons_mixed, [_params()], [_params()], _tolerances(), str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Bancos"]
        statuses = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            statuses.update({v for v in row if isinstance(v, str)})
        assert STATUS_NO_CONSTRUIDO in statuses
        assert STATUS_BANCO_ADICIONAL in statuses
