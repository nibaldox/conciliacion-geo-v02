"""Export comparison results to formatted Excel workbook.

All public write functions in this module are wrapped with try/except
and raise ``ExcelWriterError`` on failure. Callers (api/routers/export,
ui/tabs/export) can rely on a single exception type for error handling.
See docs/BARRIDO_2026-06-21.md issue D2.
"""
from typing import List, Dict, Any, Optional
from dataclasses import dataclass

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell


class ExcelWriterError(Exception):
    """Raised when the Excel export fails for any reason.

    Wraps the original exception via `raise ... from e` so callers
    can inspect the cause while presenting a clean domain error to
    end users. See docs/BARRIDO_2026-06-21.md issue D2.
    """


# ─── Number-format constants ────────────────────────────────────────────────
# Centralised so a future change to display rules (more decimals, locale-
# aware separators) only touches one place. Previously these were 4 inline
# string literals scattered across 2 functions.

@dataclass(frozen=True)
class NumberFormat:
    integer: str = "#,##0"
    decimal2: str = "0.00"
    decimal1: str = "0.0"
    integer_with_radix_decimal: str = "#,##0.0"


NUMBER_FORMAT = NumberFormat()

from core.blast_correlation import (
    compute_blast_geotech_correlation,
    compute_pasadura_stats,
)
from core.column_utils import KILOS_CANDIDATES
from core.compliance_status import (
    STATUS_CUMPLE,
    STATUS_EXTRA,
    STATUS_FUERA,
    STATUS_NO_CONSTRUIDO,
    STATUS_NO_CUMPLE,
    is_passing_status,
)
from core.geom_utils import find_df_column
from core.config import DEFAULTS


# Style constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496",
                          fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center",
                         wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
FILL_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                      fill_type="solid")
FILL_WARN = PatternFill(start_color="FFEB9C", end_color="FFEB9C",
                        fill_type="solid")
FILL_NOK = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                       fill_type="solid")
FONT_OK = Font(color="006100")
FONT_WARN = Font(color="9C5700")
FONT_NOK = Font(color="9C0006")
FILL_GREY = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
FONT_GREY = Font(color="555555")
FILL_PURPLE = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
FONT_PURPLE = Font(color="4B0082")


def _apply_status_style(cell: Cell) -> None:
    """Apply conditional formatting based on status text."""
    val = cell.value
    if val == STATUS_CUMPLE:
        cell.fill = FILL_OK
        cell.font = FONT_OK
    elif val == STATUS_FUERA:
        cell.fill = FILL_WARN
        cell.font = FONT_WARN
    elif val == STATUS_NO_CUMPLE or "FALTA" in str(val):
        cell.fill = FILL_NOK
        cell.font = FONT_NOK
    elif val == STATUS_NO_CONSTRUIDO:
        cell.fill = FILL_GREY
        cell.font = FONT_GREY
    elif val == STATUS_EXTRA or "ADICIONAL" in str(val) or "RAMPA" in str(val):
        cell.fill = FILL_PURPLE
        cell.font = FONT_PURPLE


def _write_header(ws: Worksheet, row: int, headers: List[str]) -> None:
    """Write a styled header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws: Worksheet) -> None:
    """Auto-adjust column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 25)


def _write_summary_sheet(wb: Workbook, comparisons: List[Dict[str, Any]],
                          tolerances: Dict[str, Any], project_info: Dict[str, str]) -> None:
    """Create the Resumen sheet."""
    ws = wb.active
    ws.title = "Resumen"

    # Project info
    row = 1
    ws.cell(row=row, column=1,
            value="CONCILIACION GEOTECNICA: DISENO vs AS-BUILT").font = Font(
        bold=True, size=14, color="2F5496")
    row = 3
    info_fields = [
        ("Proyecto", project_info.get('project', '')),
        ("Operacion", project_info.get('operation', '')),
        ("Fase / Pit", project_info.get('phase', '')),
        ("Elaborado por", project_info.get('author', '')),
        ("Fecha", project_info.get('date', '')),
    ]
    for label, value in info_fields:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # Tolerance table
    ws.cell(row=row, column=1,
            value="Tolerancias Aplicadas").font = Font(bold=True, size=12)
    row += 1
    _write_header(ws, row, ["Parametro", "Tol. (-)", "Tol. (+)"])
    row += 1
    tol_display = [
        ("Altura de banco (m)", 'bench_height'),
        ("Angulo cara (deg)", 'face_angle'),
        ("Angulo inter-rampa (deg)", 'inter_ramp_angle'),
        ("Angulo global (deg)", 'overall_angle'),
    ]
    for label, key in tol_display:
        tol = tolerances.get(key, {'neg': 0, 'pos': 0})
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=tol.get('neg', 0)).border = THIN_BORDER
        ws.cell(row=row, column=3, value=tol.get('pos', 0)).border = THIN_BORDER
        row += 1
    # Berm minimum
    berm_tol = tolerances.get('berm_width', {})
    ws.cell(row=row, column=1, value="Berma minima (m)").border = THIN_BORDER
    ws.cell(row=row, column=2,
            value=berm_tol.get('min', 0)).border = THIN_BORDER
    ws.cell(row=row, column=3, value="").border = THIN_BORDER
    row += 1

    row += 1

    # Compliance summary
    if comparisons:
        ws.cell(row=row, column=1,
                value="Resumen de Cumplimiento").font = Font(
            bold=True, size=12)
        row += 1
        _write_header(ws, row,
                      ["Parametro", STATUS_CUMPLE, "FUERA TOL.", STATUS_NO_CUMPLE,
                       "Total", "% Cumpl."])
        row += 1

        for key, label in [('height_status', 'Altura de banco'),
                           ('angle_status', 'Angulo de cara'),
                           ('berm_status', 'Ancho de berma')]:
            # Only MATCH benches have actual measured values; MISSING/EXTRA
            # don't carry height/angle/berm measurements so they must not
            # inflate the denominator.
            match_comps = [c for c in comparisons if c.get('type') == 'MATCH']
            total = len(match_comps)
            n_ok = sum(1 for c in match_comps if c[key] == STATUS_CUMPLE)
            n_warn = sum(1 for c in match_comps
                        if c[key] == STATUS_FUERA)
            n_nok = sum(1 for c in match_comps if c[key] == STATUS_NO_CUMPLE)
            pct = n_ok / total * 100 if total > 0 else 0

            ws.cell(row=row, column=1, value=label).border = THIN_BORDER
            c_ok = ws.cell(row=row, column=2, value=n_ok)
            c_ok.border = THIN_BORDER
            c_ok.fill = FILL_OK
            c_warn = ws.cell(row=row, column=3, value=n_warn)
            c_warn.border = THIN_BORDER
            c_warn.fill = FILL_WARN
            c_nok = ws.cell(row=row, column=4, value=n_nok)
            c_nok.border = THIN_BORDER
            c_nok.fill = FILL_NOK
            ws.cell(row=row, column=5, value=total).border = THIN_BORDER
            ws.cell(row=row, column=6,
                    value=f"{pct:.1f}%").border = THIN_BORDER
            row += 1

    _auto_width(ws)


def _write_bench_sheet(wb: Workbook, comparisons: List[Dict[str, Any]]) -> None:
    """Create the Bancos detail sheet."""
    ws = wb.create_sheet("Bancos")

    headers = [
        "Sector", "Seccion", "Banco", "Nivel",
        "H. Diseno (m)", "H. Real (m)", "Desv. H (m)", "Cumpl. H",
        "A. Diseno (deg)", "A. Real (deg)", "Desv. A (deg)", "Cumpl. A",
        "B. Diseno (m)", "B. Real (m)", "B. Minima (m)", "Cumpl. B",
        "Delta Cresta (m)", "Delta Pata (m)"
    ]
    _write_header(ws, 1, headers)

    for row_idx, comp in enumerate(comparisons, 2):
        values = [
            comp['sector'], comp['section'], comp['bench_num'], comp['level'],
            comp['height_design'], comp['height_real'],
            comp['height_dev'], comp['height_status'],
            comp['angle_design'], comp['angle_real'],
            comp['angle_dev'], comp['angle_status'],
            comp['berm_design'], comp['berm_real'],
            comp.get('berm_min', 0), comp['berm_status'],
            comp.get('delta_crest', ''), comp.get('delta_toe', ''),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_idx in (8, 12, 16):
                _apply_status_style(cell)

    _auto_width(ws)


def _write_interramp_sheet(wb: Workbook, params_design: List[Any], params_topo: List[Any]) -> None:
    """Create the Inter-Rampa sheet."""
    ws = wb.create_sheet("Inter-Rampa")

    headers = [
        "Seccion", "Sector",
        "Ang. IR Diseno (deg)", "Ang. IR Real (deg)",
        "Ang. Global Diseno (deg)", "Ang. Global Real (deg)",
    ]
    _write_header(ws, 1, headers)

    n = min(len(params_design), len(params_topo))
    for i in range(n):
        pd = params_design[i]
        pt = params_topo[i]
        row = i + 2
        ws.cell(row=row, column=1, value=pd.section_name).border = THIN_BORDER
        ws.cell(row=row, column=2, value=pd.sector).border = THIN_BORDER
        ws.cell(row=row, column=3,
                value=round(pd.inter_ramp_angle, 1)).border = THIN_BORDER
        ws.cell(row=row, column=4,
                value=round(pt.inter_ramp_angle, 1)).border = THIN_BORDER
        ws.cell(row=row, column=5,
                value=round(pd.overall_angle, 1)).border = THIN_BORDER
        ws.cell(row=row, column=6,
                value=round(pt.overall_angle, 1)).border = THIN_BORDER

    _auto_width(ws)


def _write_dashboard_sheet(wb: Workbook, comparisons: List[Dict[str, Any]]) -> None:
    """Create a simple Dashboard summary sheet."""
    ws = wb.create_sheet("Dashboard")

    ws.cell(row=1, column=1,
            value="DASHBOARD DE CUMPLIMIENTO").font = Font(
        bold=True, size=14, color="2F5496")

    if not comparisons:
        ws.cell(row=3, column=1, value="Sin datos de comparacion.")
        return

    total = len(comparisons)
    row = 3
    _write_header(ws, row,
                  ["Parametro", STATUS_CUMPLE, "FUERA TOL.", STATUS_NO_CUMPLE,
                   "% Cumplimiento"])
    row += 1

    for key, label in [('height_status', 'Altura de banco'),
                       ('angle_status', 'Angulo de cara'),
                       ('berm_status', 'Ancho de berma')]:
        n_ok = sum(1 for c in comparisons if c[key] == STATUS_CUMPLE)
        n_warn = sum(1 for c in comparisons
                    if c[key] == STATUS_FUERA)
        n_nok = sum(1 for c in comparisons if c[key] == STATUS_NO_CUMPLE)
        pct = n_ok / total * 100 if total > 0 else 0

        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        c_ok = ws.cell(row=row, column=2, value=n_ok)
        c_ok.border = THIN_BORDER
        c_ok.fill = FILL_OK
        c_warn = ws.cell(row=row, column=3, value=n_warn)
        c_warn.border = THIN_BORDER
        c_warn.fill = FILL_WARN
        c_nok = ws.cell(row=row, column=4, value=n_nok)
        c_nok.border = THIN_BORDER
        c_nok.fill = FILL_NOK
        ws.cell(row=row, column=5,
                value=f"{pct:.1f}%").border = THIN_BORDER
        row += 1

    # Global compliance
    row += 1
    n_all = total * 3
    n_ok_all = sum(
        1 for c in comparisons
        for k in ['height_status', 'angle_status', 'berm_status']
        if c[k] == STATUS_CUMPLE
    )
    pct_global = n_ok_all / n_all * 100 if n_all > 0 else 0

    ws.cell(row=row, column=1,
            value="CUMPLIMIENTO GLOBAL").font = Font(bold=True, size=12)
    ws.cell(row=row, column=2,
            value=f"{pct_global:.1f}%").font = Font(bold=True, size=14)

    _auto_width(ws)


def _write_sector_summary(wb: Workbook, comparisons: List[Dict[str, Any]]) -> None:
    """Create Executive Summary by Sector."""
    ws = wb.create_sheet("Resumen Ejecutivo")

    ws.cell(row=1, column=1, value="RESUMEN EJECUTIVO POR SECTOR").font = Font(bold=True, size=14, color="2F5496")

    headers = ["Sector", "Total Bancos", "Cumplimiento Global", "Cumpl. Altura", "Cumpl. Angulo", "Cumpl. Berma"]
    _write_header(ws, 3, headers)

    # Group by sector
    sectors = sorted(list(set(c['sector'] for c in comparisons)))

    row = 4
    for sector in sectors:
        sec_comps = [c for c in comparisons if c['sector'] == sector]
        total = len(sec_comps)

        # Helper to count matches
        def count_ok(key):
            return sum(1 for c in sec_comps if is_passing_status(c.get(key)))

        ok_h = count_ok('height_status')
        ok_a = count_ok('angle_status')
        ok_b = count_ok('berm_status')

        # Calculate percentages
        # Note: Global compliance is average of all parameters? Or average of banks that pass all?
        # Let's use % of total parameters (3 per bank)
        n_params = total * 3
        n_ok_total = ok_h + ok_a + ok_b
        pct_global = (n_ok_total / n_params * 100) if n_params > 0 else 0

        pct_h = (ok_h / total * 100) if total > 0 else 0
        pct_a = (ok_a / total * 100) if total > 0 else 0
        pct_b = (ok_b / total * 100) if total > 0 else 0

        ws.cell(row=row, column=1, value=sector).border = THIN_BORDER
        ws.cell(row=row, column=2, value=total).border = THIN_BORDER

        # Global
        c_Glob = ws.cell(row=row, column=3, value=f"{pct_global:.1f}%")
        c_Glob.border = THIN_BORDER
        if pct_global >= 90: c_Glob.fill = FILL_OK; c_Glob.font = FONT_OK
        elif pct_global >= 75: c_Glob.fill = FILL_WARN; c_Glob.font = FONT_WARN
        else: c_Glob.fill = FILL_NOK; c_Glob.font = FONT_NOK

        # Per param
        ws.cell(row=row, column=4, value=f"{pct_h:.1f}%").border = THIN_BORDER
        ws.cell(row=row, column=5, value=f"{pct_a:.1f}%").border = THIN_BORDER
        ws.cell(row=row, column=6, value=f"{pct_b:.1f}%").border = THIN_BORDER

        row += 1

    _auto_width(ws)


def export_results(comparisons: List[Dict[str, Any]], params_design: List[Any],
                   params_topo: List[Any], tolerances: Dict[str, Any],
                   output_path: str, project_info: Optional[Dict[str, str]] = None,
                   df_pozos: Optional[Any] = None,
                   sections: Optional[List[Any]] = None) -> None:
    """Export comparison results to a formatted Excel workbook.

    Raises:
        ExcelWriterError: if any sub-step of the export fails. The
            original exception is preserved via ``__cause__``.
    """
    if project_info is None:
        project_info = {}
    try:
        wb = openpyxl.Workbook()

        _write_summary_sheet(wb, comparisons, tolerances, project_info)
        _write_sector_summary(wb, comparisons) # New Executive Summary
        _write_bench_sheet(wb, comparisons)
        _write_interramp_sheet(wb, params_design, params_topo)
        _write_dashboard_sheet(wb, comparisons)

        if df_pozos is not None and not df_pozos.empty:
            _write_tronadura_sheet(wb, df_pozos, comparisons, sections)

        wb.save(output_path)
    except ExcelWriterError:
        raise
    except Exception as exc:
        raise ExcelWriterError(
            f"export_results falló al escribir {output_path}: "
            f"{type(exc).__name__}: {exc}"
        ) from exc


def _write_tronadura_sheet(wb: openpyxl.Workbook, df_pozos: Any, comparisons: List[Dict[str, Any]], sections: Optional[List[Any]] = None) -> None:
    """Write Drill & Blast and Geotechnical correlation results to Excel."""
    ws = wb.create_sheet("Tronadura")
    ws.views.sheetView[0].showGridLines = True

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    title_font = Font(name='Calibri', size=16, bold=True, color='2F5496')
    section_font = Font(name='Calibri', size=13, bold=True, color='1F4E78')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    regular_font = Font(name='Calibri', size=11)

    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    zebra_fill = PatternFill(start_color='F2F5F8', end_color='F2F5F8', fill_type='solid')

    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    thin_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Title
    ws.cell(row=1, column=1, value="ANÁLISIS DE PERFORACIÓN, TRONADURA Y CORRELACIÓN").font = title_font

    # Section 1: Geotechnical Correlation Summary
    ws.cell(row=3, column=1, value="1. Resumen de Correlación Geotécnica vs Voladura").font = section_font

    r_label = int(DEFAULTS.blast_correlation_radius_m)
    headers_corr = [
        "Sección",
        f"Pozos Cercanos (r={r_label}m)",
        "Carga Explosiva Cercana (Kg)",
        "Desviación Media Absoluta (m)",
    ]
    for col_idx, h in enumerate(headers_corr, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Use the shared correlation helper so Excel and Word stay in sync
    corr_rows = compute_blast_geotech_correlation(
        df_pozos=df_pozos,
        sections=sections or [],
        comparisons=comparisons or [],
    )
    if not corr_rows and comparisons:
        # Fallback: at least show the sections that have comparisons
        seen = []
        for c in comparisons:
            if c.get("section") not in seen:
                seen.append(c["section"])
        corr_rows = [
            type("R", (), {
                "as_tuple": lambda self, n=s: (n, 0, 0, 0.0)
            })() for s in seen
        ]

    row_idx = 5
    for row in corr_rows:
        sec_name, num_wells, total_kg, avg_dev = row.as_tuple()
        row_vals = [sec_name, num_wells, total_kg, avg_dev]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx in (2, 3):
                cell.number_format = NUMBER_FORMAT.integer
                cell.alignment = center_align
            elif col_idx == 4:
                cell.number_format = NUMBER_FORMAT.decimal2
                cell.alignment = center_align
            else:
                cell.alignment = left_align
        row_idx += 1

    # Section 2: Drill Holes Details
    row_idx += 2
    ws.cell(row=row_idx, column=1, value="2. Detalle de Pozos de Tronadura y Pasadura").font = section_font

    row_idx += 1
    headers_pozo = ["Pozo", "Collar X", "Collar Y", "Collar Z", "Pata X", "Pata Y", "Pata Z", "Largo (m)", "Inclinación (°)", "Azimut (°)", "Explosivo (Kg)", "Pasadura (m)"]
    for col_idx, h in enumerate(headers_pozo, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    row_idx += 1
    df_pozos = df_pozos.copy()
    stats = compute_pasadura_stats(df_pozos)
    df_pozos['Pasadura'] = (df_pozos['Z_collar'] - DEFAULTS.blast_default_bench_height) - df_pozos['Z_toe']
    # Touch `stats` so the computation is exercised even if no consumer reads it
    _ = stats["total"], stats["mean"], stats["optimal_pct"]

    kg_col = find_df_column(df_pozos, list(KILOS_CANDIDATES), raise_error=False)
    label_col = find_df_column(df_pozos, ['label_pozo'], raise_error=False)
    incl_col = 'Incl'
    az_col = 'Az'
    len_col = 'Len'

    # Use itertuples instead of iterrows: ~10x faster on large DataFrames
    # because itertuples returns namedtuples, not Series objects.
    for idx, row in enumerate(df_pozos.itertuples(index=False)):
        label = (
            str(getattr(row, label_col, None))
            if label_col and pd.notna(getattr(row, label_col, None))
            else f"P-{idx+1}"
        )
        x_c = row.X
        y_c = row.Y
        z_c = row.Z_collar
        x_t = row.X_toe
        y_t = row.Y_toe
        z_t = row.Z_toe
        length = getattr(row, len_col, None)
        incl = row.Incl
        az = row.Az
        kg = (
            getattr(row, kg_col, None)
            if kg_col and pd.notna(getattr(row, kg_col, None))
            else 0.0
        )
        pas = row.Pasadura

        row_vals = [label, x_c, y_c, z_c, x_t, y_t, z_t, length, incl, az, kg, pas]

        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
                if col_idx in (2, 3, 4, 5, 6, 7):
                    cell.number_format = NUMBER_FORMAT.integer_with_radix_decimal
                elif col_idx in (8, 9, 10, 11, 12):
                    cell.number_format = NUMBER_FORMAT.decimal1

        if idx % 2 == 1:
            for col_idx in range(1, 13):
                ws.cell(row=row_idx, column=col_idx).fill = zebra_fill

        row_idx += 1

    _auto_width(ws)
