"""Structured blast-processing diagnostics export (integración §3.7/4.7).

This module exports the full structured processor result produced by
``procesar_pozos`` to a multi-sheet Excel workbook and a one-page PDF
summary. Every sheet is independent and re-openable by openpyxl/pandas
in tests — no opaque single-string dump.

Sheets emitted (all optional; only non-empty ones are written):

    Pozos_Aceptados        — one row per accepted blast hole
    Filas_Rechazadas       — one row per (row × error) rejection record
    Advertencias           — event-level warnings
    Errores_Bloqueantes    — blocking errors (zero accepted, etc.)
    Resumen_Procesamiento  — counts + geometry configuration echo
    Configuracion_Geometrica — versioned contract dump
    Diagnostico_Espacial   — domain / Voronoi diagnostics

A standalone rejected-only export is provided via
:func:`export_rejected_rows_excel` so an operator can audit a failed
upload without re-running the full pipeline.
"""
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.geometry_contract import GEOMETRY_CONFIGURATION_VERSION


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
ERROR_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def _styled_header(ws, headers: List[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = None
    # Auto-width (best-effort, capped)
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 200), min_col=col_idx, max_col=col_idx):
            for c in row:
                if c.value is not None:
                    max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def _normalize_cell(value: Any) -> Any:
    """Normalize a single cell value for openpyxl.

    Integración §5.10 — openpyxl raises ``ValueError: Cannot convert
    dict to Excel`` when it receives a container (dict/list/tuple).
    We serialize any non-scalar through ``json.dumps`` with stable
    settings, so nested diagnostics round-trip safely and read back as
    JSON strings the operator can parse later. ``None`` stays empty.
    """
    if value is None:
        return None
    # Native scalars openpyxl accepts directly.
    if isinstance(value, (bool, int, float, str)):
        # NaN/inf floats are not JSON-compliant; surface as None.
        if isinstance(value, float):
            try:
                import math
                if not math.isfinite(value):
                    return None
            except (TypeError, ValueError):
                return None
        return value
    # Dates / datetimes are Excel-compatible.
    if isinstance(value, (datetime,)):
        return value
    # Containers and any other type → stable JSON.
    try:
        return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    except (TypeError, ValueError):
        return str(value)


def _write_df_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=name[:31])  # Excel sheet-name cap = 31 chars
    if df.empty:
        ws.cell(row=1, column=1, value="(sin registros)")
        return
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, record in enumerate(df.to_dict(orient="records"), start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = record[header]
            # Preserve the version field as a string so Excel does not
            # coerce "2.0" → 2.0 (integración §3.3 — single version
            # source of truth must round-trip verbatim).
            if header == "geometry_configuration_version" and value is not None:
                value = str(value)
            # Make every value JSON-safe (numpy/pandas scalars → native).
            if isinstance(value, (pd.Timestamp, datetime)):
                value = value.isoformat()
            ws.cell(row=row_idx, column=col_idx, value=_normalize_cell(value))
    _styled_header(ws, headers)


def _write_kv_sheet(wb: Workbook, name: str, kv: Dict[str, Any]) -> None:
    ws = wb.create_sheet(title=name[:31])
    ws.cell(row=1, column=1, value="clave").font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=2, value="valor").font = HEADER_FONT
    ws.cell(row=1, column=2).fill = HEADER_FILL
    for idx, (key, value) in enumerate(kv.items(), start=2):
        ws.cell(row=idx, column=1, value=str(key))
        ws.cell(row=idx, column=2, value=_normalize_cell(value))
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 80


def export_processing_diagnostics_excel(
    payload: Dict[str, Any],
    output_path: str | Path,
) -> Path:
    """Write the structured blast-processing result to a multi-sheet xlsx.

    ``payload`` is the dict produced by ``api/routers/blast.py::
    _build_upload_payload`` (or its persisted equivalent). Every field is
    optional; the workbook always contains at least the processing summary.

    Returns the output path. The file is always re-openable by openpyxl.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Drop the default sheet — we'll create our own.
    wb.remove(wb.active)

    accepted_rows = payload.get("accepted_rows") or payload.get("records") or []
    rejected_rows = payload.get("rejected_rows") or []
    event_warnings = payload.get("event_warnings") or []
    blocking_errors = payload.get("blocking_errors") or []
    summary = payload.get("processing_summary") or {}
    config = payload.get("geometry_configuration") or {}
    spatial = payload.get("spatial_diagnostics") or {}

    _write_df_sheet(wb, "Pozos_Aceptados", pd.DataFrame(accepted_rows))
    _write_df_sheet(wb, "Filas_Rechazadas", pd.DataFrame(rejected_rows))
    _write_df_sheet(wb, "Advertencias", pd.DataFrame(event_warnings))
    _write_df_sheet(wb, "Errores_Bloqueantes", pd.DataFrame(blocking_errors))
    _write_kv_sheet(wb, "Resumen_Procesamiento", summary)
    _write_kv_sheet(wb, "Configuracion_Geometrica", config)
    _write_kv_sheet(wb, "Diagnostico_Espacial", spatial)

    wb.save(str(output_path))
    return output_path


def export_rejected_rows_excel(
    rejected_rows: List[Dict[str, Any]],
    output_path: str | Path,
    *,
    config: Optional[Dict[str, Any]] = None,
) -> Path:
    """Standalone rejected-only export (independent audit artifact).

    Writes one ``Filas_Rechazadas`` sheet plus a small metadata sheet
    echoing the geometry configuration when provided. The output is a
    valid xlsx re-openable by openpyxl/pandas/Excel.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    _write_df_sheet(wb, "Filas_Rechazadas", pd.DataFrame(rejected_rows))
    if config is not None:
        meta = {
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "geometry_configuration_version": config.get(
                "geometry_configuration_version", GEOMETRY_CONFIGURATION_VERSION
            ),
            "n_rejected_rows": len(rejected_rows),
        }
        _write_kv_sheet(wb, "Metadata", meta)

    wb.save(str(output_path))
    return output_path


def read_back_excel(path: str | Path) -> Dict[str, pd.DataFrame]:
    """Re-open an exported workbook for verification (used by tests).

    Returns a dict keyed by sheet name; each value is a DataFrame whose
    first row is treated as the header. The default empty-sheet sentinel
    ``(sin registros)`` produces an empty DataFrame.
    """
    path = Path(path)
    out: Dict[str, pd.DataFrame] = {}
    xl = pd.ExcelFile(str(path), engine="openpyxl")
    for sheet in xl.sheet_names:
        df = xl.parse(sheet_name=sheet)
        # Drop the placeholder row the exporter writes for empty sheets.
        if list(df.columns) == [0] or (len(df.columns) == 1 and df.empty):
            df = pd.DataFrame()
        out[sheet] = df
    return out
