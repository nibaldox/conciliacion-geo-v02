"""Excel export for the simulation result.

Multi-sheet workbook consumable by mining engineers:

* ``Resumen``        — processing summary (energy totals, conservation,
                       temporal status, mode).
* ``Configuración``  — full SimulationConfiguration dict.
* ``Fuentes``        — source summary (counts, segments, voxels).
* ``Advertencias``   — structured warnings.
* ``Diagnósticos``   — spatial + temporal diagnostics.
* ``Mapa_Planta``    — plan slice catalogue (one row per elevation).
* ``Secciones``      — section slice catalogue.
* ``Procedencia``    — provenance + hashes.

Nested containers are serialised through ``_normalize_cell`` so the
workbook is always re-readable by openpyxl (integración §5.10).
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook

from core.blast_export import _normalize_cell, _styled_header, _write_kv_sheet, HEADER_FONT, HEADER_FILL
from core.blast_simulation.contracts import SimulationResult


def _write_df_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=name[:31])
    if df.empty:
        ws.cell(row=1, column=1, value="(sin registros)")
        return
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, record in enumerate(df.to_dict(orient="records"), start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = record[header]
            ws.cell(row=row_idx, column=col_idx, value=_normalize_cell(value))
    _styled_header(ws, headers)


def export_simulation_xlsx(
    result: SimulationResult,
    output_path: str | Path,
) -> Path:
    """Write the canonical SimulationResult to a multi-sheet xlsx.

    Returns the output path. The file is always re-openable by openpyxl.
    """
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    # 1. Resumen
    _write_kv_sheet(wb, "Resumen", result.processing_summary.to_dict())

    # 2. Configuración
    _write_kv_sheet(wb, "Configuración", result.configuration)

    # 3. Fuentes
    _write_kv_sheet(wb, "Fuentes", result.source_summary.to_dict())

    # 4. Advertencias
    warnings_df = pd.DataFrame(list(result.warnings))
    _write_df_sheet(wb, "Advertencias", warnings_df)

    # 5. Bloqueos (blocking errors)
    blocking_df = pd.DataFrame(list(result.blocking_errors))
    _write_df_sheet(wb, "Bloqueos", blocking_df)

    # 6. Diagnósticos
    diag = {}
    diag.update(result.spatial_diagnostics or {})
    diag.update({f"temporal_{k}": v for k, v in (result.temporal_diagnostics or {}).items()})
    _write_kv_sheet(wb, "Diagnósticos", diag)

    # 7. Mapa_Planta
    if result.plan_slices:
        plans_df = pd.DataFrame([s.to_dict() for s in result.plan_slices])
    else:
        plans_df = pd.DataFrame()
    _write_df_sheet(wb, "Mapa_Planta", plans_df)

    # 8. Secciones
    if result.section_slices:
        secs_df = pd.DataFrame([s.to_dict() for s in result.section_slices])
    else:
        secs_df = pd.DataFrame()
    _write_df_sheet(wb, "Secciones", secs_df)

    # 9. Procedencia
    _write_kv_sheet(wb, "Procedencia", result.provenance.to_dict())

    wb.save(out)
    return out


def read_back_simulation_xlsx(path: str | Path) -> dict[str, pd.DataFrame]:
    """Reopen the workbook produced by :func:`export_simulation_xlsx`.

    Returns a dict ``{sheet_name: DataFrame}``. Used by tests to verify
    the round-trip (every nested structure must survive as parseable
    JSON via :func:`json.loads`).
    """
    from openpyxl import load_workbook

    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(p)
    wb = load_workbook(p, read_only=True, data_only=True)
    out: dict[str, pd.DataFrame] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            out[sheet_name] = pd.DataFrame()
            continue
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        data = rows[1:]
        out[sheet_name] = pd.DataFrame(data, columns=headers)
    wb.close()
    return out
